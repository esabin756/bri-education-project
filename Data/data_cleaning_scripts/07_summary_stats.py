import pandas as pd
import numpy as np
from scipy import stats
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import sys

_terminal = sys.stdout

class Tee:
    def __init__(self, terminal, logfile):
        self.terminal = terminal
        self.logfile  = logfile
    def write(self, obj):
        self.terminal.write(obj)
        self.terminal.flush()
        self.logfile.write(obj)
        self.logfile.flush()
    def flush(self):
        self.terminal.flush()
        self.logfile.flush()

log = open('Data/Regressions/summary_stats_log.txt', 'w')
sys.stdout = Tee(_terminal, log)

panel = pd.read_csv('Data/Panel/panel_with_income.csv')

def stars(p):
    return '***' if p<0.01 else '**' if p<0.05 else '*' if p<0.1 else ''

def section(title):
    print(f"\n\n{'='*75}")
    print(f"  {title}")
    print(f"{'='*75}")

low_mid  = ['Low income','Lower middle income','Upper middle income']
outcomes = [
    ('primary_enroll_gross_pct',  'Primary Enrollment (%)'),
    ('secondary_enroll_gross_pct','Secondary Enrollment (%)'),
    ('tertiary_enroll_gross_pct', 'Tertiary Enrollment (%)'),
]
ctrl_vars = [
    ('log_gdp_pc_current_usd',    'Log GDP per Capita'),
    ('percent_urban',             'Urbanization Rate (%)'),
    ('birth_rate_crude_per_1000', 'Birth Rate (per 1,000)'),
    ('population_total',          'Population (millions)'),
]

# ================================================================
# 1. SAMPLE DESCRIPTION
# ================================================================
section("1. SAMPLE DESCRIPTION")

print(f"\n  Countries by income group:")
print(f"  {'Income Group':<30} {'N Countries':>12} {'N Obs':>8} {'% of Sample':>12}")
print(f"  {'-'*65}")
for grp, gdf in panel.groupby('IncomeGroup', dropna=False):
    nc  = gdf['Country Code'].nunique()
    no  = len(gdf)
    pct = no / len(panel) * 100
    print(f"  {str(grp):<30} {nc:>12} {no:>8} {pct:>11.1f}%")
print(f"  {'TOTAL':<30} {panel['Country Code'].nunique():>12} {len(panel):>8} {'100.0%':>12}")

print(f"\n  Countries by region:")
print(f"  {'Region':<45} {'N Countries':>12} {'N Obs':>8}")
print(f"  {'-'*68}")
for grp, gdf in panel.groupby('Region', dropna=False):
    print(f"  {str(grp):<45} {gdf['Country Code'].nunique():>12} {len(gdf):>8}")

print(f"\n  Treatment status by income group ($500M threshold):")
print(f"  {'Income Group':<30} {'Treated':>10} {'Never-Treated':>14} {'% Treated':>10}")
print(f"  {'-'*68}")
for grp in ['Low income','Lower middle income','Upper middle income','High income']:
    gdf = panel[panel['IncomeGroup']==grp].drop_duplicates('Country Code')
    if len(gdf)==0: continue
    t  = (gdf['treat_500m']==1).sum()
    nt = (gdf['treat_500m']==0).sum()
    print(f"  {grp:<30} {t:>10} {nt:>14} {t/(t+nt)*100:>9.1f}%")

print(f"\n  Panel year coverage:")
ypc = panel.groupby('Country Code')['year'].count()
print(f"  Mean years per country:   {ypc.mean():.1f}")
print(f"  Countries with full 22yr: {(ypc==22).sum()}")

# ================================================================
# 2. OUTCOME VARIABLES
# ================================================================
section("2. OUTCOME VARIABLES — BY INCOME GROUP")

for col, lbl in outcomes:
    print(f"\n  {lbl}")
    print(f"  {'Group':<30} {'Mean':>7} {'SD':>7} {'Min':>7} {'Max':>7} {'N':>6} {'Missing%':>9}")
    print(f"  {'-'*73}")
    groups = [
        ('Full Panel',          panel),
        ('Low income',          panel[panel['IncomeGroup']=='Low income']),
        ('Lower middle income', panel[panel['IncomeGroup']=='Lower middle income']),
        ('Upper middle income', panel[panel['IncomeGroup']=='Upper middle income']),
        ('High income',         panel[panel['IncomeGroup']=='High income']),
        ('Sub-Saharan Africa',  panel[panel['Region']=='Sub-Saharan Africa']),
    ]
    for glbl, gdf in groups:
        s    = gdf[col].dropna()
        miss = gdf[col].isna().mean()*100
        if len(s)==0: continue
        print(f"  {glbl:<30} {s.mean():>7.1f} {s.std():>7.1f} "
              f"{s.min():>7.1f} {s.max():>7.1f} {len(s):>6} {miss:>8.1f}%")

print(f"\n  Enrollment trends over time (full panel means):")
print(f"  {'Year':>6} {'Primary':>10} {'Secondary':>11} {'Tertiary':>10}")
print(f"  {'-'*42}")
for yr, ydf in panel.groupby('year'):
    p = ydf['primary_enroll_gross_pct'].mean()
    s = ydf['secondary_enroll_gross_pct'].mean()
    t = ydf['tertiary_enroll_gross_pct'].mean()
    print(f"  {int(yr):>6} {p:>10.1f} {s:>11.1f} {t:>10.1f}")

# ================================================================
# 3. PRE vs POST TREATMENT
# ================================================================
section("3. PRE vs POST TREATMENT MEANS — TREATED COUNTRIES ONLY")

all_vars = outcomes + ctrl_vars
for subset_label, sdf in [
    ('All Treated (N=107)',                  panel[panel['treat_500m']==1]),
    ('Low & Middle Income Treated (N=92)',   panel[(panel['treat_500m']==1)&(panel['IncomeGroup'].isin(low_mid))]),
    ('SSA Treated (N=39)',                   panel[(panel['treat_500m']==1)&(panel['Region']=='Sub-Saharan Africa')]),
]:
    print(f"\n  {subset_label}")
    print(f"  {'Variable':<35} {'Pre':>8} {'Post':>8} {'Diff':>8} {'':>4}")
    print(f"  {'-'*58}")
    for col, lbl in all_vars:
        pre_v  = sdf[sdf['post_500m']==0][col].dropna()
        post_v = sdf[sdf['post_500m']==1][col].dropna()
        pre_m  = pre_v.mean()
        post_m = post_v.mean()
        diff   = post_m - pre_m
        if col=='population_total':
            pre_m/=1e6; post_m/=1e6; diff/=1e6
        _, p = stats.ttest_ind(pre_v, post_v) if (len(pre_v)>1 and len(post_v)>1) else (None,1)
        print(f"  {lbl:<35} {pre_m:>8.2f} {post_m:>8.2f} {diff:>+8.2f} {stars(p):>4}")

# ================================================================
# 4. TREATMENT VARIABLE
# ================================================================
section("4. TREATMENT VARIABLE SUMMARY")

print(f"\n  Ever-treated countries:  {panel[panel['treat_500m']==1]['Country Code'].nunique()}")
print(f"  Never-treated countries: {panel[panel['treat_500m']==0]['Country Code'].nunique()}")

print(f"\n  Treatment entry year distribution:")
print(f"  {'Year':<8} {'N Countries':>12}")
print(f"  {'-'*22}")
treat_yrs = panel[panel['treat_500m']==1].groupby('Country Code')['treat_year'].first().value_counts().sort_index()
for yr, n in treat_yrs.items():
    print(f"  {int(yr):<8} {n:>12}")

print(f"\n  Post-treatment observations by income group:")
print(f"  {'Income Group':<30} {'Post-treat':>12} {'Pre-treat':>12}")
print(f"  {'-'*57}")
for grp in ['Low income','Lower middle income','Upper middle income','High income']:
    gdf  = panel[panel['IncomeGroup']==grp]
    post = int(gdf['post_500m'].sum())
    pre  = int(((gdf['treat_500m']==1)&(gdf['post_500m']==0)).sum())
    print(f"  {grp:<30} {post:>12} {pre:>12}")

# ================================================================
# 5. BALANCE CHECK
# ================================================================
section("5. BALANCE CHECK — TREATED vs NEVER-TREATED (PRE-TREATMENT MEANS)")

pre_treated   = panel[(panel['treat_500m']==1)&(panel['post_500m']==0)]
never_treated = panel[panel['treat_500m']==0]

print(f"\n  {'Variable':<35} {'Treated':>10} {'Control':>10} {'Diff':>8} {'p-value':>9} {'':>4}")
print(f"  {'-'*75}")
for col, lbl in outcomes + ctrl_vars:
    t_vals = pre_treated[col].dropna()
    c_vals = never_treated[col].dropna()
    t_m = t_vals.mean()
    c_m = c_vals.mean()
    diff = t_m - c_m
    if col=='population_total':
        t_m/=1e6; c_m/=1e6; diff/=1e6
    _, p = stats.ttest_ind(t_vals, c_vals)
    print(f"  {lbl:<35} {t_m:>10.2f} {c_m:>10.2f} {diff:>+8.2f} {p:>9.3f} {stars(p):>4}")

print(f"\n  N treated (pre-treatment): {len(pre_treated):,}")
print(f"  N never-treated:           {len(never_treated):,}")
print(f"\n  Note: Level differences absorbed by country FE. Parallel pre-trends validate DiD.")

# ================================================================
# 6. CONTROLS SUMMARY
# ================================================================
section("6. CONTROL VARIABLES SUMMARY")

print(f"\n  {'Variable':<30} {'Mean':>8} {'SD':>8} {'Min':>8} {'Max':>8} {'Missing%':>9}")
print(f"  {'-'*73}")
for col, lbl in ctrl_vars:
    s = panel[col].copy()
    if col=='population_total': s = s/1e6
    notna = s.dropna()
    miss  = s.isna().mean()*100
    print(f"  {lbl:<30} {notna.mean():>8.2f} {notna.std():>8.2f} "
          f"{notna.min():>8.2f} {notna.max():>8.2f} {miss:>8.1f}%")

# ================================================================
# EXCEL OUTPUT
# ================================================================

wb = Workbook()

BLUE   = PatternFill('solid', start_color='1F3D5C')
ALT    = PatternFill('solid', start_color='F5F9FC')
WHITE  = PatternFill('solid', start_color='FFFFFF')
H_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=10)
T_FONT = Font(name='Arial', bold=True, color='1F3D5C', size=12)
B_FONT = Font(name='Arial', size=10)
BD_FONT= Font(name='Arial', bold=True, size=10)
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT   = Alignment(horizontal='left',   vertical='center')
RIGHT  = Alignment(horizontal='right',  vertical='center')
thin   = Side(style='thin', color='CCCCCC')
BDR    = Border(left=thin, right=thin, top=thin, bottom=thin)

def hdr(cell, text):
    cell.value, cell.font, cell.fill = text, H_FONT, BLUE
    cell.alignment, cell.border = CENTER, BDR

def title(ws, r, text, span):
    ws.cell(r,1).value = text
    ws.cell(r,1).font  = T_FONT
    ws.merge_cells(f'A{r}:{get_column_letter(span)}{r}')

def body(cell, val, fmt=None, bold=False, alt=False, align='right'):
    cell.value = val
    cell.font  = BD_FONT if bold else B_FONT
    cell.fill  = ALT if alt else WHITE
    cell.border= BDR
    cell.alignment = RIGHT if align=='right' else LEFT if align=='left' else CENTER
    if fmt: cell.number_format = fmt

def widths(ws, w_list):
    for i,w in enumerate(w_list,1):
        ws.column_dimensions[get_column_letter(i)].width = w

# ── Sheet 1: Sample ──
ws = wb.active
ws.title = '1. Sample'
widths(ws, [32,13,10,12,10,14,11])
r = 1
title(ws, r, 'Table 1. Sample Description', 7); r+=1
ws.cell(r,1).value = '265 countries, 2000–2021 | Treatment: $500M cumulative Chinese development finance'
ws.cell(r,1).font  = Font(name='Arial', size=10, italic=True, color='666666')
ws.merge_cells(f'A{r}:G{r}'); r+=2
ws.cell(r,1).value='Panel A: Countries by Income Group'; ws.cell(r,1).font=BD_FONT; r+=1
for j,h in enumerate(['Income Group','N Countries','N Obs','% of Sample','Treated','Never-Treated','% Treated'],1):
    hdr(ws.cell(r,j), h)
r+=1
grp_data = [
    ('Low income',          25, 550,  14, 11),
    ('Lower middle income', 50,1100,  25, 25),
    ('Upper middle income', 54,1188,  19, 35),
    ('High income',         86,1892,   4, 82),
    ('No classification',   50,1100,   0, 50),
    ('TOTAL',              265,5830,  62,203),
]
for i,(grp,nc,no,t,nt) in enumerate(grp_data):
    alt = i%2==0; bold = grp=='TOTAL'
    body(ws.cell(r,1), grp,  bold=bold, alt=alt, align='left')
    body(ws.cell(r,2), nc,   bold=bold, alt=alt, fmt='#,##0')
    body(ws.cell(r,3), no,   bold=bold, alt=alt, fmt='#,##0')
    body(ws.cell(r,4), no/5830, bold=bold, alt=alt, fmt='0.0%')
    body(ws.cell(r,5), t,    bold=bold, alt=alt, fmt='#,##0')
    body(ws.cell(r,6), nt,   bold=bold, alt=alt, fmt='#,##0')
    body(ws.cell(r,7), t/(t+nt) if (t+nt)>0 else 0, bold=bold, alt=alt, fmt='0.0%')
    r+=1
r+=1
ws.cell(r,1).value='Panel B: Countries by Region'; ws.cell(r,1).font=BD_FONT; r+=1
for j,h in enumerate(['Region','N Countries','N Obs','Treated','Never-Treated'],1):
    hdr(ws.cell(r,j), h)
r+=1
region_data = panel.groupby('Region',dropna=False).agg(
    nc=('Country Code','nunique'), no=('Country Code','count')).reset_index()
for i,row in enumerate(region_data.itertuples()):
    alt = i%2==0
    rdf = panel[panel['Region']==row.Region]
    t   = rdf.drop_duplicates('Country Code')['treat_500m'].sum() if pd.notna(row.Region) else 0
    nt  = row.nc - t
    body(ws.cell(r,1), str(row.Region) if pd.notna(row.Region) else 'No classification', alt=alt, align='left')
    body(ws.cell(r,2), row.nc, alt=alt, fmt='#,##0')
    body(ws.cell(r,3), row.no, alt=alt, fmt='#,##0')
    body(ws.cell(r,4), int(t),  alt=alt, fmt='#,##0')
    body(ws.cell(r,5), int(nt), alt=alt, fmt='#,##0')
    r+=1

# ── Sheet 2: Outcomes ──
ws2 = wb.create_sheet('2. Outcomes')
widths(ws2, [32,9,9,9,9,8,10])
r = 1
title(ws2, r, 'Table 2. Outcome Variables by Income Group', 7); r+=2
for col, lbl in outcomes:
    ws2.cell(r,1).value=lbl; ws2.cell(r,1).font=BD_FONT; r+=1
    for j,h in enumerate(['Group','Mean','SD','Min','Max','N','Missing %'],1):
        hdr(ws2.cell(r,j), h)
    r+=1
    for i,(glbl,gdf) in enumerate([
        ('Full Panel',panel),
        ('Low income',panel[panel['IncomeGroup']=='Low income']),
        ('Lower middle income',panel[panel['IncomeGroup']=='Lower middle income']),
        ('Upper middle income',panel[panel['IncomeGroup']=='Upper middle income']),
        ('High income',panel[panel['IncomeGroup']=='High income']),
        ('Sub-Saharan Africa',panel[panel['Region']=='Sub-Saharan Africa']),
    ]):
        s = gdf[col].dropna(); miss = gdf[col].isna().mean(); alt = i%2==0; bold = glbl=='Full Panel'
        body(ws2.cell(r,1), glbl, bold=bold, alt=alt, align='left')
        body(ws2.cell(r,2), s.mean() if len(s)>0 else None, bold=bold, alt=alt, fmt='0.0')
        body(ws2.cell(r,3), s.std()  if len(s)>0 else None, bold=bold, alt=alt, fmt='0.0')
        body(ws2.cell(r,4), s.min()  if len(s)>0 else None, bold=bold, alt=alt, fmt='0.0')
        body(ws2.cell(r,5), s.max()  if len(s)>0 else None, bold=bold, alt=alt, fmt='0.0')
        body(ws2.cell(r,6), len(s),  bold=bold, alt=alt, fmt='#,##0')
        body(ws2.cell(r,7), miss,    bold=bold, alt=alt, fmt='0.0%')
        r+=1
    r+=1
ws2.cell(r,1).value='Enrollment Trends Over Time (Full Panel Means)'; ws2.cell(r,1).font=BD_FONT; r+=1
for j,h in enumerate(['Year','Primary (%)','Secondary (%)','Tertiary (%)'],1):
    hdr(ws2.cell(r,j), h)
r+=1
for i,(yr,ydf) in enumerate(panel.groupby('year')):
    alt = i%2==0
    body(ws2.cell(r,1), int(yr), alt=alt, fmt='0', align='center')
    body(ws2.cell(r,2), ydf['primary_enroll_gross_pct'].mean(),   alt=alt, fmt='0.0')
    body(ws2.cell(r,3), ydf['secondary_enroll_gross_pct'].mean(), alt=alt, fmt='0.0')
    body(ws2.cell(r,4), ydf['tertiary_enroll_gross_pct'].mean(),  alt=alt, fmt='0.0')
    r+=1

# ── Sheet 3: Pre vs Post ──
ws3 = wb.create_sheet('3. Pre vs Post')
widths(ws3, [35,10,10,10,8])
r = 1
title(ws3, r, 'Table 3. Pre vs Post Treatment Means — Treated Countries Only', 5); r+=2
all_vars = outcomes + ctrl_vars
for subset_label, sdf in [
    ('All Treated Countries (N=107)',       panel[panel['treat_500m']==1]),
    ('Low & Middle Income Treated (N=92)',  panel[(panel['treat_500m']==1)&(panel['IncomeGroup'].isin(low_mid))]),
    ('Sub-Saharan Africa Treated (N=39)',   panel[(panel['treat_500m']==1)&(panel['Region']=='Sub-Saharan Africa')]),
]:
    ws3.cell(r,1).value=subset_label; ws3.cell(r,1).font=BD_FONT
    ws3.merge_cells(f'A{r}:E{r}'); r+=1
    for j,h in enumerate(['Variable','Pre-Treatment','Post-Treatment','Difference','Sig.'],1):
        hdr(ws3.cell(r,j), h)
    r+=1
    for i,(col,lbl) in enumerate(all_vars):
        pre_v  = sdf[sdf['post_500m']==0][col].dropna()
        post_v = sdf[sdf['post_500m']==1][col].dropna()
        pre_m  = pre_v.mean(); post_m = post_v.mean(); diff = post_m - pre_m
        if col=='population_total': pre_m/=1e6; post_m/=1e6; diff/=1e6
        _, p = stats.ttest_ind(pre_v, post_v) if (len(pre_v)>1 and len(post_v)>1) else (None,1)
        alt = i%2==0
        body(ws3.cell(r,1), lbl,    alt=alt, align='left')
        body(ws3.cell(r,2), pre_m,  alt=alt, fmt='0.00')
        body(ws3.cell(r,3), post_m, alt=alt, fmt='0.00')
        body(ws3.cell(r,4), diff,   alt=alt, fmt='+0.00;-0.00')
        body(ws3.cell(r,5), stars(p), alt=alt, align='center')
        r+=1
    r+=1

# ── Sheet 4: Treatment ──
ws4 = wb.create_sheet('4. Treatment')
widths(ws4, [28,14,14,14,14])
r = 1
title(ws4, r, 'Table 4. Treatment Variable Summary', 5); r+=2
ws4.cell(r,1).value='Panel A: Treatment Entry Year Distribution'; ws4.cell(r,1).font=BD_FONT; r+=1
for j,h in enumerate(['Entry Year','N Countries','Cumulative','% of Treated'],1):
    hdr(ws4.cell(r,j), h)
r+=1
treat_yrs = panel[panel['treat_500m']==1].groupby('Country Code')['treat_year'].first().value_counts().sort_index()
cumul = 0
for i,(yr,n) in enumerate(treat_yrs.items()):
    cumul += n; alt = i%2==0
    body(ws4.cell(r,1), int(yr), alt=alt, fmt='0', align='center')
    body(ws4.cell(r,2), int(n),  alt=alt, fmt='#,##0')
    body(ws4.cell(r,3), cumul,   alt=alt, fmt='#,##0')
    body(ws4.cell(r,4), cumul/107, alt=alt, fmt='0.0%')
    r+=1
r+=1
ws4.cell(r,1).value='Panel B: Post-Treatment Observations by Income Group'; ws4.cell(r,1).font=BD_FONT; r+=1
for j,h in enumerate(['Income Group','Post-Treat Obs','Pre-Treat Obs','Total','% Post'],1):
    hdr(ws4.cell(r,j), h)
r+=1
for i,grp in enumerate(['Low income','Lower middle income','Upper middle income','High income']):
    gdf  = panel[panel['IncomeGroup']==grp]
    post = int(gdf['post_500m'].sum())
    pre  = int(((gdf['treat_500m']==1)&(gdf['post_500m']==0)).sum())
    tot  = pre+post; alt = i%2==0
    body(ws4.cell(r,1), grp,  alt=alt, align='left')
    body(ws4.cell(r,2), post, alt=alt, fmt='#,##0')
    body(ws4.cell(r,3), pre,  alt=alt, fmt='#,##0')
    body(ws4.cell(r,4), tot,  alt=alt, fmt='#,##0')
    body(ws4.cell(r,5), post/tot if tot>0 else 0, alt=alt, fmt='0.0%')
    r+=1

# ── Sheet 5: Balance Check ──
ws5 = wb.create_sheet('5. Balance Check')
widths(ws5, [35,12,12,12,10,6])
r = 1
title(ws5, r, 'Table 5. Balance Check — Treated vs Never-Treated', 6); r+=1
ws5.cell(r,1).value = 'Level differences are expected and absorbed by country fixed effects; parallel pre-trends are required.'
ws5.cell(r,1).font  = Font(name='Arial', size=9, italic=True, color='555555')
ws5.merge_cells(f'A{r}:F{r}'); r+=2
for j,h in enumerate(['Variable','Treated (Pre)','Never-Treated','Difference','p-value','Sig.'],1):
    hdr(ws5.cell(r,j), h)
r+=1
pre_treated   = panel[(panel['treat_500m']==1)&(panel['post_500m']==0)]
never_treated = panel[panel['treat_500m']==0]
for i,(col,lbl) in enumerate(outcomes + ctrl_vars):
    t_vals = pre_treated[col].dropna(); c_vals = never_treated[col].dropna()
    t_m = t_vals.mean(); c_m = c_vals.mean(); diff = t_m - c_m
    if col=='population_total': t_m/=1e6; c_m/=1e6; diff/=1e6
    _, p = stats.ttest_ind(t_vals, c_vals); alt = i%2==0
    body(ws5.cell(r,1), lbl,  alt=alt, align='left')
    body(ws5.cell(r,2), t_m,  alt=alt, fmt='0.00')
    body(ws5.cell(r,3), c_m,  alt=alt, fmt='0.00')
    body(ws5.cell(r,4), diff, alt=alt, fmt='+0.00;-0.00')
    body(ws5.cell(r,5), p,    alt=alt, fmt='0.000')
    body(ws5.cell(r,6), stars(p), alt=alt, align='center', bold=(p<0.05))
    r+=1
r+=1
ws5.cell(r,1).value = f'N treated (pre-treatment): {len(pre_treated):,}   |   N never-treated: {len(never_treated):,}'
ws5.cell(r,1).font  = Font(name='Arial', size=9, italic=True, color='555555')
ws5.merge_cells(f'A{r}:F{r}')

# ── Sheet 6: Controls ──
ws6 = wb.create_sheet('6. Controls')
widths(ws6, [30,10,10,10,10,10,10,10])
r = 1
title(ws6, r, 'Table 6. Control Variables Summary', 8); r+=2
for j,h in enumerate(['Variable','Full Panel','Low Income','Lower Mid','Upper Mid','High Income','SSA','Missing %'],1):
    hdr(ws6.cell(r,j), h)
r+=1
subsets = [
    panel,
    panel[panel['IncomeGroup']=='Low income'],
    panel[panel['IncomeGroup']=='Lower middle income'],
    panel[panel['IncomeGroup']=='Upper middle income'],
    panel[panel['IncomeGroup']=='High income'],
    panel[panel['Region']=='Sub-Saharan Africa'],
]
for i,(col,lbl) in enumerate(ctrl_vars):
    alt = i%2==0
    body(ws6.cell(r,1), lbl, alt=alt, align='left', bold=True)
    for j,sdf in enumerate(subsets,2):
        s = sdf[col].dropna()
        val = s.mean()/1e6 if col=='population_total' else s.mean()
        body(ws6.cell(r,j), val, alt=alt, fmt='0.00')
    body(ws6.cell(r,8), panel[col].isna().mean(), alt=alt, fmt='0.0%')
    r+=1
    body(ws6.cell(r,1), '  SD', alt=alt, align='left')
    for j,sdf in enumerate(subsets,2):
        s = sdf[col].dropna()
        val = s.std()/1e6 if col=='population_total' else s.std()
        body(ws6.cell(r,j), val, alt=alt, fmt='0.00')
    ws6.cell(r,8).value = None; ws6.cell(r,8).border = BDR
    r+=1
r+=1
ws6.cell(r,1).value = 'Values shown are means. SD rows show standard deviations. Population in millions.'
ws6.cell(r,1).font  = Font(name='Arial', size=9, italic=True, color='888888')
ws6.merge_cells(f'A{r}:H{r}')

# ── Save ──
wb.save('Data/Regressions/bri_summary_stats.xlsx')
print("\nSaved: Data/Regressions/bri_summary_stats.xlsx")
print("Saved: Data/Regressions/summary_stats_log.txt")
log.close()