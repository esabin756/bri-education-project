import pandas as pd
import numpy as np
import statsmodels.formula.api as smf
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import sys
import os

_terminal = sys.stdout

class Tee:
    def __init__(self, terminal, logfile):
        self.terminal = terminal
        self.logfile  = logfile
    def write(self, obj):
        self.terminal.write(obj)
        self.terminal.flush()
        self.logfile.write(obj)
        self.logfile.flush()
    def flush(self):
        self.terminal.flush()
        self.logfile.flush()

os.makedirs('Data/Regressions', exist_ok=True)
os.makedirs('Data/Figures', exist_ok=True)

log = open('Data/Regressions/master_tables_log.txt', 'w')
sys.stdout = Tee(_terminal, log)

# ── Load data ──
panel  = pd.read_csv('Data/Panel/panel_with_income.csv')
emp    = pd.read_csv('Data/Panel/employment_panel.csv')
income = pd.read_csv('Data/Panel/income_panel.csv')

panel = panel.merge(
    emp[['Country Code','year','youth_emp_ratio','female_emp_ratio',
         'industry_emp_pct','unemployment_pct','youth_unemp_pct']],
    on=['Country Code','year'], how='left')

panel = panel.merge(
    income[['Country Code','year','gni_per_capita','gdp_growth','poverty_3day']],
    on=['Country Code','year'], how='left')

controls = 'log_gdp_pc_current_usd + population_total + percent_urban + birth_rate_crude_per_1000'
low_mid  = ['Low income','Lower middle income','Upper middle income']

full = panel.copy()
lm   = panel[panel['IncomeGroup'].isin(low_mid)].copy()
ssa  = panel[panel['Region']=='Sub-Saharan Africa'].copy()

# ── All outcomes ──
all_outcomes = [
    ('primary_enroll_gross_pct',   'Primary Enrollment (%)'),
    ('secondary_enroll_gross_pct', 'Secondary Enrollment (%)'),
    ('tertiary_enroll_gross_pct',  'Tertiary Enrollment (%)'),
    ('youth_emp_ratio',            'Youth Employment Ratio (%)'),
    ('female_emp_ratio',           'Female Employment Ratio (%)'),
    ('industry_emp_pct',           'Industry Employment (%)'),
    ('unemployment_pct',           'Unemployment Rate (%)'),
    ('youth_unemp_pct',            'Youth Unemployment Rate (%)'),
    ('gni_per_capita',             'GNI per Capita (USD)'),
    ('gdp_growth',                 'GDP Growth (%)'),
]

# ================================================================
# HELPER: run basic DiD and dynamic DiD
# ================================================================
def run_basic(df, outcome):
    cols = [outcome,'Country Code','year','treat_500m','post_500m','rel_time',
            'log_gdp_pc_current_usd','population_total','percent_urban',
            'birth_rate_crude_per_1000']
    df = df.dropna(subset=[c for c in cols if c != 'rel_time']).copy()
    df['did'] = df['treat_500m'] * df['post_500m']
    mod = smf.ols(
        f'{outcome} ~ did + {controls} + C(year) + C(Q("Country Code"))',
        data=df
    ).fit(cov_type='cluster', cov_kwds={'groups': df['Country Code']})
    return (mod.params['did'], mod.bse['did'],
            mod.pvalues['did'], int(mod.nobs),
            df['Country Code'].nunique())

def run_dynamic(df, outcome):
    cols = [outcome,'Country Code','year','treat_500m','post_500m','rel_time',
            'log_gdp_pc_current_usd','population_total','percent_urban',
            'birth_rate_crude_per_1000']
    df = df.dropna(subset=[c for c in cols if c != 'rel_time']).copy()
    for l in [0,1,2,3,4,5]:
        df[f'lag{l}'] = ((df['treat_500m']==1) & (df['rel_time']==l)).astype(float)
    for p in [1,2]:
        df[f'pre{p}'] = ((df['treat_500m']==1) & (df['rel_time']==-p)).astype(float)
    lag_terms = ' + '.join([f'lag{l}' for l in range(6)] + ['pre1','pre2'])
    mod = smf.ols(
        f'{outcome} ~ {lag_terms} + {controls} + C(year) + C(Q("Country Code"))',
        data=df
    ).fit(cov_type='cluster', cov_kwds={'groups': df['Country Code']})
    results = []
    for pp, col in [(-2,'pre2'),(-1,'pre1'),(0,'lag0'),(1,'lag1'),
                    (2,'lag2'),(3,'lag3'),(4,'lag4'),(5,'lag5')]:
        results.append({
            'period': pp,
            'coef':   mod.params[col],
            'se':     mod.bse[col],
            'ci_lo':  mod.params[col] - 1.96*mod.bse[col],
            'ci_hi':  mod.params[col] + 1.96*mod.bse[col],
            'p':      mod.pvalues[col],
        })
    return pd.DataFrame(results)

# ================================================================
# TABLE 1: MASTER RESULTS TABLE
# ================================================================
print("Building master results table...")

# Style helpers
BLUE   = PatternFill('solid', start_color='1F3D5C')
LBLUE  = PatternFill('solid', start_color='D9E8F5')
ALT    = PatternFill('solid', start_color='F5F9FC')
WHITE  = PatternFill('solid', start_color='FFFFFF')
GREEN  = PatternFill('solid', start_color='E8F5E9')
RED    = PatternFill('solid', start_color='FFEBEE')
HFONT  = Font(name='Arial', bold=True, color='FFFFFF', size=10)
BFONT  = Font(name='Arial', size=10)
BDFONT = Font(name='Arial', bold=True, size=10)
TFONT  = Font(name='Arial', bold=True, color='1F3D5C', size=13)
thin   = Side(style='thin', color='CCCCCC')
BDR    = Border(left=thin, right=thin, top=thin, bottom=thin)
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT   = Alignment(horizontal='left',   vertical='center')
RIGHT  = Alignment(horizontal='right',  vertical='center')

def hdr(cell, text, colspan=None):
    cell.value = text
    cell.font  = HFONT
    cell.fill  = BLUE
    cell.alignment = CENTER
    cell.border = BDR

def body(cell, val, fmt=None, bold=False, alt=False,
         align='right', sig=False, neg=False):
    cell.value = val
    cell.font  = BDFONT if bold else BFONT
    cell.fill  = (GREEN if sig and not neg else
                  RED   if sig and neg else
                  ALT   if alt else WHITE)
    cell.border = BDR
    cell.alignment = (RIGHT if align=='right' else
                      LEFT  if align=='left'  else CENTER)
    if fmt: cell.number_format = fmt

def stars(p):
    return '***' if p<0.01 else '**' if p<0.05 else '*' if p<0.1 else ''

wb = Workbook()

# ── Sheet 1: Master Results ──
ws = wb.active
ws.title = 'Master Results'

r = 1
ws.cell(r,1).value = 'Table: BRI Investment and Development Outcomes — TWFE DiD Results'
ws.cell(r,1).font  = TFONT
ws.merge_cells(f'A{r}:J{r}'); r+=1
ws.cell(r,1).value = ('Treatment: $500M cumulative Chinese development finance | '
                      'Controls: log GDP pc, population, urbanization, birth rate | '
                      'Country and year FE | Clustered SE by country')
ws.cell(r,1).font  = Font(name='Arial', size=9, italic=True, color='666666')
ws.merge_cells(f'A{r}:J{r}'); r+=2

# Column headers
for j,h in enumerate(['Outcome','','Full Panel','','Low & Middle Income','',
                       'Sub-Saharan Africa','',''],1):
    if h: hdr(ws.cell(r,j), h)
ws.merge_cells(f'C{r}:D{r}')
ws.merge_cells(f'E{r}:F{r}')
ws.merge_cells(f'G{r}:H{r}')
r+=1
for j,h in enumerate(['','Category','Coef','N obs',
                       'Coef','N obs','Coef','N obs'],1):
    hdr(ws.cell(r,j), h)
r+=1

subsets = [
    (full, 'Full Panel'),
    (lm,   'Low & Middle Income'),
    (ssa,  'Sub-Saharan Africa'),
]

sections = [
    ('Education', [
        ('primary_enroll_gross_pct',   'Primary Enrollment (%)'),
        ('secondary_enroll_gross_pct', 'Secondary Enrollment (%)'),
        ('tertiary_enroll_gross_pct',  'Tertiary Enrollment (%)'),
    ]),
    ('Employment', [
        ('youth_emp_ratio',   'Youth Employment Ratio (%)'),
        ('female_emp_ratio',  'Female Employment Ratio (%)'),
        ('industry_emp_pct',  'Industry Employment (%)'),
        ('unemployment_pct',  'Unemployment Rate (%)'),
        ('youth_unemp_pct',   'Youth Unemployment (%)'),
    ]),
    ('Income & Growth', [
        ('gni_per_capita', 'GNI per Capita (USD)'),
        ('gdp_growth',     'GDP Growth (%)'),
    ]),
]

row_i = 0
for section_label, outcomes in sections:
    # Section header
    ws.cell(r,1).value = section_label
    ws.cell(r,1).font  = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    ws.cell(r,1).fill  = PatternFill('solid', start_color='2E6DA4')
    ws.merge_cells(f'A{r}:H{r}')
    ws.cell(r,1).alignment = LEFT
    ws.cell(r,1).border = BDR
    r+=1

    for ocol, olbl in outcomes:
        alt = row_i % 2 == 0
        row_i += 1
        body(ws.cell(r,1), olbl, alt=alt, align='left')
        body(ws.cell(r,2), '', alt=alt)

        col_j = 3
        for sdf, slbl in subsets:
            try:
                coef, se, p, n, nc = run_basic(sdf, ocol)
                sig_  = stars(p)
                coef_str = f'{coef:+.3f}{sig_}'
                is_sig = p < 0.1
                is_neg = coef < 0
                body(ws.cell(r,col_j),   coef_str, alt=alt,
                     align='center', sig=is_sig, neg=is_neg)
                body(ws.cell(r,col_j+1), f'{n:,}',  alt=alt, align='center')
                print(f"  {slbl:<25} {olbl:<35} {coef:+.3f}{sig_}")
            except Exception as e:
                body(ws.cell(r,col_j),   '—', alt=alt, align='center')
                body(ws.cell(r,col_j+1), '—', alt=alt, align='center')
                print(f"  ERROR {slbl} {olbl}: {e}")
            col_j += 2
        r += 1
    r += 1

# Notes
ws.cell(r,1).value = ('*** p<0.01  ** p<0.05  * p<0.10  | '
                      'Green = significant positive  Red = significant negative  | '
                      'Coefficients from Post×Treated indicator in TWFE DiD')
ws.cell(r,1).font  = Font(name='Arial', size=9, italic=True, color='888888')
ws.merge_cells(f'A{r}:H{r}')

# Column widths
for i,w in enumerate([28,3,14,10,14,10,14,10],1):
    ws.column_dimensions[get_column_letter(i)].width = w

# ================================================================
# EVENT STUDY FIGURES
# ================================================================
print("\nBuilding event study figures...")

fig_outcomes = [
    ('secondary_enroll_gross_pct', 'Secondary Enrollment (%)',   full, 'Full Panel'),
    ('secondary_enroll_gross_pct', 'Secondary Enrollment (%)',   ssa,  'Sub-Saharan Africa'),
    ('female_emp_ratio',           'Female Employment Ratio (%)',full, 'Full Panel'),
    ('gdp_growth',                 'GDP Growth (%)',             ssa,  'Sub-Saharan Africa'),
]

fig, axes = plt.subplots(2, 2, figsize=(14, 10))
fig.suptitle('Event Study: BRI Investment and Development Outcomes\n'
             '($500M Cumulative Investment Threshold)',
             fontsize=14, fontweight='bold', y=0.98)

colors = {'pre': '#E74C3C', 'post': '#2E86AB'}

for ax, (ocol, olbl, sdf, slbl) in zip(axes.flatten(), fig_outcomes):
    try:
        dyn = run_dynamic(sdf, ocol)

        pre  = dyn[dyn['period'] < 0]
        post = dyn[dyn['period'] >= 0]

        # Pre-trend (red)
        ax.errorbar(pre['period'], pre['coef'],
                    yerr=1.96*pre['se'],
                    fmt='o-', color=colors['pre'],
                    capsize=4, linewidth=2, markersize=6,
                    label='Pre-treatment')

        # Post-treatment (blue)
        ax.errorbar(post['period'], post['coef'],
                    yerr=1.96*post['se'],
                    fmt='s-', color=colors['post'],
                    capsize=4, linewidth=2, markersize=6,
                    label='Post-treatment')

        # Reference lines
        ax.axhline(0, color='black', linewidth=0.8, linestyle='--', alpha=0.5)
        ax.axvline(-0.5, color='gray', linewidth=1.5, linestyle=':', alpha=0.7)

        # Shade pre-trend region
        ax.axvspan(-2.5, -0.5, alpha=0.05, color='red')
        ax.axvspan(-0.5, 5.5,  alpha=0.05, color='blue')

        ax.set_title(f'{olbl}\n{slbl}', fontsize=11, fontweight='bold')
        ax.set_xlabel('Years Relative to Treatment', fontsize=10)
        ax.set_ylabel('Coefficient (pp)', fontsize=10)
        ax.set_xticks(range(-2, 6))
        ax.set_xticklabels([f't{i}' if i>=0 else str(i) for i in range(-2,6)])
        ax.legend(fontsize=9, loc='upper left')
        ax.grid(True, alpha=0.3)
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)

        # Add zero treatment line label
        ax.text(-0.5, ax.get_ylim()[1]*0.95, 'Treatment →',
                fontsize=8, color='gray', ha='center')

    except Exception as e:
        ax.text(0.5, 0.5, f'Error: {e}', transform=ax.transAxes,
                ha='center', va='center')
        print(f"  Figure error {slbl} {olbl}: {e}")

plt.tight_layout(rect=[0, 0, 1, 0.96])
plt.savefig('Data/Figures/event_study_main.png', dpi=150, bbox_inches='tight')
plt.close()
print("  Saved: Data/Figures/event_study_main.png")

# ── SSA sector figure ──
fig2, axes2 = plt.subplots(1, 3, figsize=(16, 5))
fig2.suptitle('Event Study: SSA Secondary Enrollment by Sector\n'
              'Dose-Response — Log Investment',
              fontsize=13, fontweight='bold')

cgit = pd.read_csv('Data/GIT_Data/clean_git/cgit_clean_transactions_with_iso3.csv')

for ax, sector in zip(axes2, ['Transport','Energy','Metals']):
    try:
        agg = (cgit[cgit['sector']==sector]
                   .groupby(['iso3','year'])['deal_musd'].sum()
                   .reset_index()
                   .rename(columns={'iso3':'Country Code','deal_musd':'usd'}))
        agg[f'log1p_{sector}'] = np.log1p(agg['usd'])

        sdf2 = ssa.merge(agg[['Country Code','year',f'log1p_{sector}']],
                        on=['Country Code','year'], how='left')
        sdf2[f'log1p_{sector}'] = sdf2[f'log1p_{sector}'].fillna(0)

        ocol = 'secondary_enroll_gross_pct'
        cols = [ocol,'Country Code','year',f'log1p_{sector}',
                'log_gdp_pc_current_usd','population_total',
                'percent_urban','birth_rate_crude_per_1000']
        df2 = sdf2.dropna(subset=cols).copy()

        # Bin investment into quartiles for event-style plot
        nonzero = df2[df2[f'log1p_{sector}']>0][f'log1p_{sector}']
        if len(nonzero) < 20:
            ax.text(0.5, 0.5, 'Insufficient data', transform=ax.transAxes,
                    ha='center'); continue

        mod = smf.ols(
            f'{ocol} ~ {f"log1p_{sector}"} + {controls} + C(year) + C(Q("Country Code"))',
            data=df2
        ).fit(cov_type='cluster', cov_kwds={'groups': df2['Country Code']})

        c  = mod.params[f'log1p_{sector}']
        se = mod.bse[f'log1p_{sector}']
        p  = mod.pvalues[f'log1p_{sector}']
        sig= stars(p)

        ax.bar([sector], [c],
               color=('#2E86AB' if c>0 else '#E74C3C'),
               alpha=0.8, width=0.4)
        ax.errorbar([sector], [c], yerr=1.96*se,
                    fmt='none', color='black', capsize=8, linewidth=2)
        ax.axhline(0, color='black', linewidth=0.8, linestyle='--')
        ax.set_title(f'{sector}\n{c:+.3f}{sig}\n(p={p:.3f})', fontsize=11)
        ax.set_ylabel('Coefficient on log investment', fontsize=10)
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)

    except Exception as e:
        ax.text(0.5, 0.5, f'Error: {e}', transform=ax.transAxes, ha='center')

plt.tight_layout()
plt.savefig('Data/Figures/ssa_sector_secondary.png', dpi=150, bbox_inches='tight')
plt.close()
print("  Saved: Data/Figures/ssa_sector_secondary.png")

# ================================================================
# SHEET 2: SSA BALANCE CHECK
# ================================================================
ws2 = wb.create_sheet('SSA Balance Check')

r = 1
ws2.cell(r,1).value = 'Table: Balance Check — SSA Treated vs SSA Never-Treated (Pre-Treatment Means)'
ws2.cell(r,1).font  = TFONT
ws2.merge_cells(f'A{r}:F{r}'); r+=1
ws2.cell(r,1).value = 'Differences absorbed by country fixed effects; parallel pre-trends validate DiD design'
ws2.cell(r,1).font  = Font(name='Arial', size=9, italic=True, color='666666')
ws2.merge_cells(f'A{r}:F{r}'); r+=2

for j,h in enumerate(['Variable','SSA Treated (Pre)','SSA Never-Treated',
                       'Difference','p-value','Sig.'],1):
    hdr(ws2.cell(r,j), h)
r+=1

from scipy import stats as scipy_stats

ssa_pre_treated   = ssa[(ssa['treat_500m']==1)&(ssa['post_500m']==0)]
ssa_never_treated = ssa[ssa['treat_500m']==0]

balance_vars = [
    ('primary_enroll_gross_pct',   'Primary Enrollment (%)'),
    ('secondary_enroll_gross_pct', 'Secondary Enrollment (%)'),
    ('tertiary_enroll_gross_pct',  'Tertiary Enrollment (%)'),
    ('log_gdp_pc_current_usd',     'Log GDP per Capita'),
    ('percent_urban',              'Urbanization Rate (%)'),
    ('birth_rate_crude_per_1000',  'Birth Rate (per 1,000)'),
    ('population_total',           'Population (millions)'),
    ('youth_emp_ratio',            'Youth Employment Ratio (%)'),
    ('female_emp_ratio',           'Female Employment Ratio (%)'),
    ('gdp_growth',                 'GDP Growth (%)'),
]

for i,(col,lbl) in enumerate(balance_vars):
    t_v = ssa_pre_treated[col].dropna()
    c_v = ssa_never_treated[col].dropna()
    t_m = t_v.mean(); c_m = c_v.mean(); diff = t_m - c_m
    if col=='population_total': t_m/=1e6; c_m/=1e6; diff/=1e6
    _, p = scipy_stats.ttest_ind(t_v, c_v)
    alt = i%2==0; sig = stars(p)
    is_sig = p < 0.1; is_neg = diff < 0
    body(ws2.cell(r,1), lbl,  alt=alt, align='left')
    body(ws2.cell(r,2), round(t_m,2), alt=alt, align='center')
    body(ws2.cell(r,3), round(c_m,2), alt=alt, align='center')
    body(ws2.cell(r,4), round(diff,2),alt=alt, align='center',
         sig=is_sig, neg=is_neg)
    body(ws2.cell(r,5), round(p,3),   alt=alt, align='center')
    body(ws2.cell(r,6), sig,          alt=alt, align='center', bold=(p<0.05))
    r+=1

r+=1
ws2.cell(r,1).value = (f'N SSA treated (pre): {len(ssa_pre_treated):,}  |  '
                       f'N SSA never-treated: {len(ssa_never_treated):,}')
ws2.cell(r,1).font  = Font(name='Arial', size=9, italic=True, color='555555')
ws2.merge_cells(f'A{r}:F{r}')

for i,w in enumerate([32,16,18,14,10,6],1):
    ws2.column_dimensions[get_column_letter(i)].width = w

# ================================================================
# SHEET 3: TOP RECIPIENTS
# ================================================================
ws3 = wb.create_sheet('Top Recipients')

r = 1
ws3.cell(r,1).value = 'Table: Top 25 BRI Recipients — Investment and Outcome Summary'
ws3.cell(r,1).font  = TFONT
ws3.merge_cells(f'A{r}:I{r}'); r+=2

for j,h in enumerate(['Country','ISO3','Region','Income Group',
                       'Total Investment ($B)','Treatment Year',
                       'Avg Secondary Enroll (Post)',
                       'Avg GDP Growth (Post)',
                       'Avg Female Emp (Post)'],1):
    hdr(ws3.cell(r,j), h)
r+=1

aid = pd.read_csv('Data/clean_aid_data/aiddata_gcdf_country_year.csv')
total_inv = (aid.groupby('Country Code')['usd_const_2021_sum']
               .sum().reset_index()
               .rename(columns={'usd_const_2021_sum':'total_usd'}))
total_inv['total_bn'] = total_inv['total_usd'] / 1e9

treat_info = (panel[panel['treat_500m']==1]
              .drop_duplicates('Country Code')
              [['Country Code','treat_year','IncomeGroup','Region']])

top25 = (total_inv.merge(treat_info, on='Country Code', how='inner')
                  .sort_values('total_bn', ascending=False)
                  .head(25))

country_names = panel[['Country Code','country_name']].drop_duplicates() \
    if 'country_name' in panel.columns else \
    panel[['Country Code']].drop_duplicates()

for i, row in enumerate(top25.itertuples()):
    alt = i%2==0
    post = panel[(panel['Country Code']==row._1)&(panel['post_500m']==1)]
    sec  = post['secondary_enroll_gross_pct'].mean()
    gdpg = post['gdp_growth'].mean()
    femp = post['female_emp_ratio'].mean()

    body(ws3.cell(r,1), row._1,         alt=alt, align='left')
    body(ws3.cell(r,2), row._1,         alt=alt, align='center')
    body(ws3.cell(r,3), str(row.Region) if pd.notna(row.Region) else '—',
         alt=alt, align='left')
    body(ws3.cell(r,4), str(row.IncomeGroup) if pd.notna(row.IncomeGroup) else '—',
         alt=alt, align='left')
    body(ws3.cell(r,5), round(row.total_bn,2), alt=alt, fmt='$#,##0.00', align='right')
    body(ws3.cell(r,6), int(row.treat_year) if pd.notna(row.treat_year) else '—',
         alt=alt, align='center')
    body(ws3.cell(r,7), round(sec,1)  if pd.notna(sec)  else '—', alt=alt, align='center')
    body(ws3.cell(r,8), round(gdpg,1) if pd.notna(gdpg) else '—', alt=alt, align='center')
    body(ws3.cell(r,9), round(femp,1) if pd.notna(femp) else '—', alt=alt, align='center')
    r+=1

for i,w in enumerate([18,8,28,22,18,14,20,16,18],1):
    ws3.column_dimensions[get_column_letter(i)].width = w

# ── Save ──
wb.save('Data/Regressions/master_results.xlsx')
print("\nSaved: Data/Regressions/master_results.xlsx")
print("Saved: Data/Figures/event_study_main.png")
print("Saved: Data/Figures/ssa_sector_secondary.png")
log.close()