import pandas as pd
import numpy as np
import statsmodels.formula.api as smf
import sys

# ── Redirect print output to file AND terminal ──
class Tee:
    def __init__(self, *files):
        self.files = files
    def write(self, obj):
        for f in self.files:
            f.write(obj)
            f.flush()
    def flush(self):
        for f in self.files:
            f.flush()

log = open('Data/Regressions/did_results_log.txt', 'w')
sys.stdout = Tee(sys.stdout, log)

panel = pd.read_csv('Data/Panel/panel_with_income.csv')

# ── Add event study dummies (now includes lags 4 and 5) ──
def add_event_vars(df, lags=[0,1,2,3,4,5], pres=[1,2]):
    df = df.copy()
    for l in lags:
        df[f'lag{l}'] = ((df['treat_500m']==1) & (df['rel_time']==l)).astype(float)
    for p in pres:
        df[f'pre{p}'] = ((df['treat_500m']==1) & (df['rel_time']==-p)).astype(float)
    return df

panel = add_event_vars(panel)
controls = 'log_gdp_pc_current_usd + population_total + percent_urban + birth_rate_crude_per_1000'

def run_did(df, outcome, label):
    cols_needed = [outcome, 'Country Code', 'year',
                   'log_gdp_pc_current_usd', 'population_total',
                   'percent_urban', 'birth_rate_crude_per_1000']
    df = df.dropna(subset=cols_needed).copy()
    df['did'] = df['treat_500m'] * df['post_500m']

    print(f"\n{'='*55}")
    print(f"OUTCOME: {label}  (N countries={df['Country Code'].nunique()})")
    print(f"{'='*55}")

    # ── Basic DiD ──
    formula_basic = f'{outcome} ~ did + {controls} + C(year) + C(Q("Country Code"))'
    mod = smf.ols(formula_basic, data=df).fit(
        cov_type='cluster', cov_kwds={'groups': df['Country Code']}
    )
    c, se, t, p = mod.params['did'], mod.bse['did'], mod.tvalues['did'], mod.pvalues['did']
    stars = '***' if p<0.01 else '**' if p<0.05 else '*' if p<0.1 else ''
    print(f"\n[1] Basic DiD — N={int(mod.nobs)}")
    print(f"    Post x Treated: {c:+.3f}{stars}  SE={se:.3f}  t={t:.2f}  p={p:.3f}")

    # ── Dynamic DiD ──
    lag_terms = ' + '.join([f'lag{l}' for l in [0,1,2,3,4,5]] + [f'pre{p}' for p in [1,2]])
    formula_dyn = f'{outcome} ~ {lag_terms} + {controls} + C(year) + C(Q("Country Code"))'
    mod_dyn = smf.ols(formula_dyn, data=df).fit(
        cov_type='cluster', cov_kwds={'groups': df['Country Code']}
    )

    print(f"\n[2] Dynamic DiD — N={int(mod_dyn.nobs)}")
    print(f"    {'Period':<10} {'Coef':>8} {'SE':>8} {'t':>7} {'p':>7} {'':>4}")
    print(f"    {'-'*45}")
    for pp, col in [(-2,'pre2'),(-1,'pre1'),(0,'lag0'),(1,'lag1'),
                    (2,'lag2'),(3,'lag3'),(4,'lag4'),(5,'lag5')]:
        c  = mod_dyn.params[col]
        se = mod_dyn.bse[col]
        t  = mod_dyn.tvalues[col]
        p  = mod_dyn.pvalues[col]
        stars = '***' if p<0.01 else '**' if p<0.05 else '*' if p<0.1 else ''
        lbl = f"Pre({abs(pp)})" if pp < 0 else f"Lag({pp})"
        print(f"    {lbl:<10} {c:>+8.3f} {se:>8.3f} {t:>7.2f} {p:>7.3f} {stars:>4}")

# ── Define subsets ──
low_mid  = ['Low income', 'Lower middle income', 'Upper middle income']
full     = panel.copy()
lm_only  = panel[panel['IncomeGroup'].isin(low_mid)].copy()
ssa      = panel[panel['Region'] == 'Sub-Saharan Africa'].copy()

# ── Top 10 within low/middle income only ──
top10 = (panel[panel['IncomeGroup'].isin(low_mid)]
               .groupby('Country Code')['usd_const_2021']
               .sum()
               .nlargest(10)
               .index.tolist())
print(f"Top 10 low/middle income by cumulative investment: {top10}")
top10_panel = panel[panel['Country Code'].isin(top10)].copy()

# ── Run all regressions ──
for subset, label in [
    (full,        'FULL PANEL'),
    (lm_only,     'LOW & MIDDLE INCOME ONLY'),
    (ssa,         'SUB-SAHARAN AFRICA'),
    (top10_panel, 'TOP 10 LOW/MIDDLE INCOME RECIPIENTS'),
]:
    print(f"\n\n{'#'*55}")
    print(f"  {label}")
    print(f"{'#'*55}")
    run_did(subset, 'primary_enroll_gross_pct',   'Primary Enrollment (%)')
    run_did(subset, 'secondary_enroll_gross_pct', 'Secondary Enrollment (%)')
    run_did(subset, 'tertiary_enroll_gross_pct',  'Tertiary Enrollment (%)')

# ── Dose-response: continuous log investment ──
def run_dose_response(df, outcome, label):
    cols_needed = [outcome, 'Country Code', 'year', 'log1p_usd_const_2021',
                   'log_gdp_pc_current_usd', 'population_total',
                   'percent_urban', 'birth_rate_crude_per_1000']
    df = df.dropna(subset=cols_needed).copy()

    formula = (f'{outcome} ~ log1p_usd_const_2021 + {controls} '
               f'+ C(year) + C(Q("Country Code"))')
    mod = smf.ols(formula, data=df).fit(
        cov_type='cluster', cov_kwds={'groups': df['Country Code']}
    )
    c  = mod.params['log1p_usd_const_2021']
    se = mod.bse['log1p_usd_const_2021']
    t  = mod.tvalues['log1p_usd_const_2021']
    p  = mod.pvalues['log1p_usd_const_2021']
    stars = '***' if p<0.01 else '**' if p<0.05 else '*' if p<0.1 else ''
    print(f"\n  {label}")
    print(f"  N={int(mod.nobs)}, Countries={df['Country Code'].nunique()}")
    print(f"  Log investment: {c:+.4f}{stars}  SE={se:.4f}  t={t:.2f}  p={p:.3f}")

print(f"\n\n{'#'*55}")
print(f"  DOSE-RESPONSE (continuous log investment)")
print(f"{'#'*55}")
for subset, label in [
    (full,    'Full Panel'),
    (lm_only, 'Low & Middle Income'),
    (ssa,     'Sub-Saharan Africa'),
]:
    print(f"\n--- {label} ---")
    run_dose_response(subset, 'primary_enroll_gross_pct',   'Primary Enrollment (%)')
    run_dose_response(subset, 'secondary_enroll_gross_pct', 'Secondary Enrollment (%)')
    run_dose_response(subset, 'tertiary_enroll_gross_pct',  'Tertiary Enrollment (%)')
    # ── FIRST DIFFERENCES ──
def run_fd(df, outcome, label):
    cols_needed = [outcome, 'Country Code', 'year', 'did',
                   'log_gdp_pc_current_usd', 'population_total',
                   'percent_urban', 'birth_rate_crude_per_1000']
    df = df.copy().sort_values(['Country Code', 'year'])
    
    # First difference the outcome and controls
    df[f'd_{outcome}'] = df.groupby('Country Code')[outcome].diff()
    df['d_log_gdp']    = df.groupby('Country Code')['log_gdp_pc_current_usd'].diff()
    df['d_urban']      = df.groupby('Country Code')['percent_urban'].diff()
    df['d_birthrate']  = df.groupby('Country Code')['birth_rate_crude_per_1000'].diff()
    df['d_pop']        = df.groupby('Country Code')['population_total'].diff()
    df['did']          = df['treat_500m'] * df['post_500m']

    cols_needed = [f'd_{outcome}', 'Country Code', 'year', 'did',
                   'd_log_gdp', 'd_urban', 'd_birthrate', 'd_pop']
    df = df.dropna(subset=cols_needed).copy()

    print(f"\n{'='*55}")
    print(f"OUTCOME: {label}  (N countries={df['Country Code'].nunique()})")
    print(f"{'='*55}")

    # Basic FD DiD
    formula = (f'd_{outcome} ~ did + d_log_gdp + d_urban + d_birthrate + d_pop'
               f' + C(year)')
    mod = smf.ols(formula, data=df).fit(
        cov_type='cluster', cov_kwds={'groups': df['Country Code']}
    )
    c, se, t, p = mod.params['did'], mod.bse['did'], mod.tvalues['did'], mod.pvalues['did']
    stars = '***' if p<0.01 else '**' if p<0.05 else '*' if p<0.1 else ''
    print(f"\n[FD Basic DiD] N={int(mod.nobs)}")
    print(f"    Post x Treated: {c:+.3f}{stars}  SE={se:.3f}  t={t:.2f}  p={p:.3f}")

    # FD with lags
    for l in [1, 2, 3]:
        df[f'did_lag{l}'] = df.groupby('Country Code')['did'].shift(l)
    
    df = df.dropna(subset=['did_lag1','did_lag2','did_lag3']).copy()
    formula_lags = (f'd_{outcome} ~ did + did_lag1 + did_lag2 + did_lag3'
                    f' + d_log_gdp + d_urban + d_birthrate + d_pop + C(year)')
    mod_l = smf.ols(formula_lags, data=df).fit(
        cov_type='cluster', cov_kwds={'groups': df['Country Code']}
    )

    print(f"\n[FD with Lags] N={int(mod_l.nobs)}")
    print(f"    {'Variable':<15} {'Coef':>8} {'SE':>8} {'t':>7} {'p':>7} {'':>4}")
    print(f"    {'-'*48}")
    for var, lbl in [('did','Post(t)'),('did_lag1','Post(t-1)'),
                     ('did_lag2','Post(t-2)'),('did_lag3','Post(t-3)')]:
        c  = mod_l.params[var]
        se = mod_l.bse[var]
        t  = mod_l.tvalues[var]
        p  = mod_l.pvalues[var]
        stars = '***' if p<0.01 else '**' if p<0.05 else '*' if p<0.1 else ''
        print(f"    {lbl:<15} {c:>+8.3f} {se:>8.3f} {t:>7.2f} {p:>7.3f} {stars:>4}")

print(f"\n\n{'#'*55}")
print(f"  FIRST DIFFERENCES")
print(f"{'#'*55}")
for subset, label in [
    (full,    'FULL PANEL'),
    (lm_only, 'LOW & MIDDLE INCOME'),
    (ssa,     'SUB-SAHARAN AFRICA'),
]:
    print(f"\n\n--- {label} ---")
    run_fd(subset, 'primary_enroll_gross_pct',   'Primary Enrollment (%)')
    run_fd(subset, 'secondary_enroll_gross_pct', 'Secondary Enrollment (%)')
    run_fd(subset, 'tertiary_enroll_gross_pct',  'Tertiary Enrollment (%)')
# ── Save regression results to Excel ──
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

BLUE  = PatternFill('solid', start_color='1F3D5C')
ALT   = PatternFill('solid', start_color='F5F9FC')
WHITE = PatternFill('solid', start_color='FFFFFF')
HFONT = Font(name='Arial', bold=True, color='FFFFFF', size=10)
BFONT = Font(name='Arial', size=10)
BDFONT= Font(name='Arial', bold=True, size=10)
TFONT = Font(name='Arial', bold=True, color='1F3D5C', size=12)
thin  = Side(style='thin', color='CCCCCC')
BDR   = Border(left=thin, right=thin, top=thin, bottom=thin)
CENTER= Alignment(horizontal='center', vertical='center')
LEFT  = Alignment(horizontal='left',   vertical='center')
RIGHT = Alignment(horizontal='right',  vertical='center')

def hdr(cell, text):
    cell.value, cell.font, cell.fill = text, HFONT, BLUE
    cell.alignment, cell.border = CENTER, BDR

def body(cell, val, fmt=None, bold=False, alt=False, align='right'):
    cell.value  = val
    cell.font   = BDFONT if bold else BFONT
    cell.fill   = ALT if alt else WHITE
    cell.border = BDR
    cell.alignment = RIGHT if align=='right' else LEFT if align=='left' else CENTER
    if fmt: cell.number_format = fmt

def set_widths(ws, w):
    for i,v in enumerate(w,1):
        ws.column_dimensions[get_column_letter(i)].width = v

def write_did_results(ws, results_list, sheet_title):
    """
    results_list: list of dicts with keys:
      subset, outcome, basic_coef, basic_se, basic_t, basic_p,
      dynamic: list of (period_label, coef, se, t, p)
      nobs, n_countries
    """
    r = 1
    ws.cell(r,1).value = sheet_title
    ws.cell(r,1).font  = TFONT
    ws.merge_cells(f'A{r}:H{r}'); r+=1
    ws.cell(r,1).value = 'Treatment: $500M cumulative Chinese development finance | TWFE DiD with clustered SEs by country'
    ws.cell(r,1).font  = Font(name='Arial', size=9, italic=True, color='666666')
    ws.merge_cells(f'A{r}:H{r}'); r+=2

    for res in results_list:
        # Subset + outcome header
        ws.cell(r,1).value = f"{res['subset']} — {res['outcome']}"
        ws.cell(r,1).font  = BDFONT
        ws.cell(r,1).fill  = PatternFill('solid', start_color='D9E8F5')
        ws.merge_cells(f'A{r}:H{r}')
        ws.cell(r,1).border = BDR; r+=1

        # Basic DiD row
        for j,h in enumerate(['Specification','N Obs','N Countries','Coefficient','Std Error','t-stat','p-value','Sig.'],1):
            hdr(ws.cell(r,j), h)
        r+=1
        p = res['basic_p']
        sig = '***' if p<0.01 else '**' if p<0.05 else '*' if p<0.1 else ''
        body(ws.cell(r,1), 'Basic DiD (Post × Treated)', align='left')
        body(ws.cell(r,2), res['nobs'],      fmt='#,##0')
        body(ws.cell(r,3), res['n_countries'],fmt='#,##0')
        body(ws.cell(r,4), res['basic_coef'],fmt='+0.000;-0.000')
        body(ws.cell(r,5), res['basic_se'],  fmt='0.000')
        body(ws.cell(r,6), res['basic_t'],   fmt='0.00')
        body(ws.cell(r,7), res['basic_p'],   fmt='0.000')
        body(ws.cell(r,8), sig, align='center', bold=(p<0.05))
        r+=2

        # Dynamic DiD
        ws.cell(r,1).value = 'Dynamic DiD — Event Study Coefficients'
        ws.cell(r,1).font  = BDFONT; r+=1
        for j,h in enumerate(['Period','','N Obs','N Countries','Coefficient','Std Error','t-stat','p-value'],1):
            hdr(ws.cell(r,j), h)
        r+=1
        for i,(period, coef, se, t, pval) in enumerate(res['dynamic']):
            is_pre = 'Pre' in period
            alt = i%2==0
            sig = '***' if pval<0.01 else '**' if pval<0.05 else '*' if pval<0.1 else ''
            body(ws.cell(r,1), period,  alt=alt, align='left',
                 bold=False)
            body(ws.cell(r,2), '← pre-trend' if is_pre else '', alt=alt, align='left')
            body(ws.cell(r,3), res['nobs'],       alt=alt, fmt='#,##0')
            body(ws.cell(r,4), res['n_countries'],alt=alt, fmt='#,##0')
            body(ws.cell(r,5), coef, alt=alt, fmt='+0.000;-0.000')
            body(ws.cell(r,6), se,   alt=alt, fmt='0.000')
            body(ws.cell(r,7), t,    alt=alt, fmt='0.00')
            body(ws.cell(r,8), pval, alt=alt, fmt='0.000')
            if sig:
                ws.cell(r,8).value = f"{pval:.3f} {sig}"
            r+=1
        r+=2

    # Note
    ws.cell(r,1).value = '*** p<0.01  ** p<0.05  * p<0.10  |  Standard errors clustered by country  |  Controls: log GDP pc, population, urbanization rate, birth rate  |  Country and year FE included'
    ws.cell(r,1).font  = Font(name='Arial', size=9, italic=True, color='888888')
    ws.merge_cells(f'A{r}:H{r}')

    set_widths(ws, [22,14,13,13,13,10,10,10])

# ── Re-run regressions and collect results for export ──
def collect_did(df, outcome, outcome_label, subset_label):
    cols_needed = [outcome, 'Country Code', 'year',
                   'log_gdp_pc_current_usd','population_total',
                   'percent_urban','birth_rate_crude_per_1000']
    df = df.dropna(subset=cols_needed).copy()
    df['did'] = df['treat_500m'] * df['post_500m']

    # Basic
    mod = smf.ols(
        f'{outcome} ~ did + {controls} + C(year) + C(Q("Country Code"))',
        data=df
    ).fit(cov_type='cluster', cov_kwds={'groups': df['Country Code']})

    # Dynamic
    lag_terms = ' + '.join([f'lag{l}' for l in [0,1,2,3,4,5]] + ['pre1','pre2'])
    mod_dyn = smf.ols(
        f'{outcome} ~ {lag_terms} + {controls} + C(year) + C(Q("Country Code"))',
        data=df
    ).fit(cov_type='cluster', cov_kwds={'groups': df['Country Code']})

    dynamic = []
    for pp, col in [(-2,'pre2'),(-1,'pre1'),(0,'lag0'),(1,'lag1'),
                    (2,'lag2'),(3,'lag3'),(4,'lag4'),(5,'lag5')]:
        lbl = f"Pre({abs(pp)})" if pp<0 else f"Lag({pp})"
        dynamic.append((lbl, mod_dyn.params[col], mod_dyn.bse[col],
                        mod_dyn.tvalues[col], mod_dyn.pvalues[col]))

    return {
        'subset':      subset_label,
        'outcome':     outcome_label,
        'basic_coef':  mod.params['did'],
        'basic_se':    mod.bse['did'],
        'basic_t':     mod.tvalues['did'],
        'basic_p':     mod.pvalues['did'],
        'nobs':        int(mod.nobs),
        'n_countries': df['Country Code'].nunique(),
        'dynamic':     dynamic,
    }

# Collect all results
all_results = []
for sdf, slbl in [
    (full,        'Full Panel'),
    (lm_only,     'Low & Middle Income'),
    (ssa,         'Sub-Saharan Africa'),
    (top10_panel, 'Top 10 Low/Mid Recipients'),
]:
    for ocol, olbl in [
        ('primary_enroll_gross_pct',   'Primary Enrollment (%)'),
        ('secondary_enroll_gross_pct', 'Secondary Enrollment (%)'),
        ('tertiary_enroll_gross_pct',  'Tertiary Enrollment (%)'),
    ]:
        try:
            all_results.append(collect_did(sdf, ocol, olbl, slbl))
        except Exception as e:
            print(f"Skipped {slbl} {olbl}: {e}")

# Write to Excel — one sheet per subset
from itertools import groupby
wb2 = Workbook()
first = True
for subset_label, grp in groupby(all_results, key=lambda x: x['subset']):
    safe_title = subset_label[:31].replace('/', '-').replace('\\', '-')
ws = wb2.active if first else wb2.create_sheet(safe_title)
if first:
     ws.title = safe_title
first = False
write_did_results(ws, list(grp), f'DiD Results — {subset_label}')

# Dose response sheet
ws_dr = wb2.create_sheet('Dose-Response')
r = 1
ws_dr.cell(r,1).value = 'Dose-Response — Continuous Log Investment'
ws_dr.cell(r,1).font  = TFONT
ws_dr.merge_cells('A1:G1'); r+=2
for j,h in enumerate(['Subset','Outcome','Coefficient','Std Error','t-stat','p-value','Sig.'],1):
    hdr(ws_dr.cell(r,j), h)
r+=1
for sdf, slbl in [(full,'Full Panel'),(lm_only,'Low & Middle Income'),(ssa,'Sub-Saharan Africa')]:
    for ocol, olbl in [
        ('primary_enroll_gross_pct',  'Primary Enrollment (%)'),
        ('secondary_enroll_gross_pct','Secondary Enrollment (%)'),
        ('tertiary_enroll_gross_pct', 'Tertiary Enrollment (%)'),
    ]:
        try:
            cols = [ocol,'Country Code','year','log1p_usd_const_2021',
                    'log_gdp_pc_current_usd','population_total',
                    'percent_urban','birth_rate_crude_per_1000']
            df2 = sdf.dropna(subset=cols).copy()
            mod = smf.ols(
                f'{ocol} ~ log1p_usd_const_2021 + {controls} + C(year) + C(Q("Country Code"))',
                data=df2
            ).fit(cov_type='cluster', cov_kwds={'groups': df2['Country Code']})
            c  = mod.params['log1p_usd_const_2021']
            se = mod.bse['log1p_usd_const_2021']
            t  = mod.tvalues['log1p_usd_const_2021']
            p  = mod.pvalues['log1p_usd_const_2021']
            sig = '***' if p<0.01 else '**' if p<0.05 else '*' if p<0.1 else ''
            alt = r%2==0
            body(ws_dr.cell(r,1), slbl,  alt=alt, align='left')
            body(ws_dr.cell(r,2), olbl,  alt=alt, align='left')
            body(ws_dr.cell(r,3), c,     alt=alt, fmt='+0.0000;-0.0000')
            body(ws_dr.cell(r,4), se,    alt=alt, fmt='0.0000')
            body(ws_dr.cell(r,5), t,     alt=alt, fmt='0.00')
            body(ws_dr.cell(r,6), p,     alt=alt, fmt='0.000')
            body(ws_dr.cell(r,7), sig,   alt=alt, align='center', bold=(p<0.05))
            r+=1
        except Exception as e:
            print(f"Skipped DR {slbl} {olbl}: {e}")
set_widths(ws_dr, [22,25,13,13,10,10,6])

wb2.save('Data/Regressions/did_results.xlsx')
print("Saved: Data/Regressions/did_results.xlsx")
log.close()
print("Saved: Data/Regressions/did_results_log.txt")