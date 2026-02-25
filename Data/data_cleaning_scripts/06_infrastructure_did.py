import pandas as pd
import numpy as np
import statsmodels.formula.api as smf

# ── Load data ──
panel = pd.read_csv('Data/Panel/panel_with_income.csv')
cgit  = pd.read_csv('Data/GIT_Data/clean_git/cgit_clean_transactions_with_iso3.csv')

countries = ['LAO','KHM','SDN','TZA','MOZ','BEN','BGD','TGO','ETH','AFG']

# ── Aggregate CGIT to country-year: infrastructure only ──
infra = cgit[cgit['sector'].isin(['Energy','Transport'])].copy()
infra_agg = (infra.groupby(['iso3','year'])
                  .agg(infra_deals=('deal_musd','count'),
                       infra_usd=('deal_musd','sum'))
                  .reset_index()
                  .rename(columns={'iso3':'Country Code'}))

infra_agg['log1p_infra_usd'] = np.log1p(infra_agg['infra_usd'])

print("Infrastructure deals by country-year:")
print(infra_agg[infra_agg['Country Code'].isin(countries)].groupby('Country Code')['infra_deals'].sum())

# ── Subset panel to 10 countries and merge ──
sub = panel[panel['Country Code'].isin(countries)].copy()
sub = sub.merge(infra_agg, on=['Country Code','year'], how='left')
sub['infra_usd']       = sub['infra_usd'].fillna(0)
sub['infra_deals']     = sub['infra_deals'].fillna(0)
sub['log1p_infra_usd'] = sub['log1p_infra_usd'].fillna(0)

print("\nPanel shape after merge:", sub.shape)
print("Country-years with infra investment:", (sub['infra_usd']>0).sum())

# ── Controls ──
controls = 'log_gdp_pc_current_usd + population_total + percent_urban + birth_rate_crude_per_1000'

# ── Dose-response: log infrastructure investment ──
def run_infra(df, outcome, label):
    cols_needed = [outcome, 'Country Code', 'year', 'log1p_infra_usd',
                   'log_gdp_pc_current_usd', 'population_total',
                   'percent_urban', 'birth_rate_crude_per_1000']
    df = df.dropna(subset=cols_needed).copy()

    formula = (f'{outcome} ~ log1p_infra_usd + {controls} '
               f'+ C(year) + C(Q("Country Code"))')
    mod = smf.ols(formula, data=df).fit(
        cov_type='cluster', cov_kwds={'groups': df['Country Code']}
    )
    c  = mod.params['log1p_infra_usd']
    se = mod.bse['log1p_infra_usd']
    t  = mod.tvalues['log1p_infra_usd']
    p  = mod.pvalues['log1p_infra_usd']
    stars = '***' if p<0.01 else '**' if p<0.05 else '*' if p<0.1 else ''
    print(f"\n  {label}")
    print(f"  N={int(mod.nobs)}, Countries={df['Country Code'].nunique()}")
    print(f"  Log infra investment: {c:+.4f}{stars}  SE={se:.4f}  t={t:.2f}  p={p:.3f}")

print(f"\n{'#'*55}")
print(f"  INFRASTRUCTURE DOSE-RESPONSE (Energy + Transport)")
print(f"  10 low-income countries")
print(f"{'#'*55}")
run_infra(sub, 'primary_enroll_gross_pct',   'Primary Enrollment (%)')
run_infra(sub, 'secondary_enroll_gross_pct', 'Secondary Enrollment (%)')

# ── Also try with lagged infrastructure investment ──
sub['log1p_infra_usd_lag1'] = sub.groupby('Country Code')['log1p_infra_usd'].shift(1)
sub['log1p_infra_usd_lag2'] = sub.groupby('Country Code')['log1p_infra_usd'].shift(2)
sub['log1p_infra_usd_lag3'] = sub.groupby('Country Code')['log1p_infra_usd'].shift(3)

def run_infra_lags(df, outcome, label):
    lag_controls = 'log1p_infra_usd + log1p_infra_usd_lag1 + log1p_infra_usd_lag2 + log1p_infra_usd_lag3'
    cols_needed = [outcome, 'Country Code', 'year',
                   'log1p_infra_usd','log1p_infra_usd_lag1',
                   'log1p_infra_usd_lag2','log1p_infra_usd_lag3',
                   'log_gdp_pc_current_usd','population_total',
                   'percent_urban','birth_rate_crude_per_1000']
    df = df.dropna(subset=cols_needed).copy()

    formula = (f'{outcome} ~ {lag_controls} + {controls} '
               f'+ C(year) + C(Q("Country Code"))')
    mod = smf.ols(formula, data=df).fit(
        cov_type='cluster', cov_kwds={'groups': df['Country Code']}
    )

    print(f"\n  {label}")
    print(f"  N={int(mod.nobs)}, Countries={df['Country Code'].nunique()}")
    print(f"  {'Variable':<25} {'Coef':>8} {'SE':>8} {'t':>7} {'p':>7}")
    print(f"  {'-'*55}")
    for var in ['log1p_infra_usd','log1p_infra_usd_lag1',
                'log1p_infra_usd_lag2','log1p_infra_usd_lag3']:
        c  = mod.params[var]
        se = mod.bse[var]
        t  = mod.tvalues[var]
        p  = mod.pvalues[var]
        stars = '***' if p<0.01 else '**' if p<0.05 else '*' if p<0.1 else ''
        print(f"  {var:<25} {c:>+8.4f} {se:>8.4f} {t:>7.2f} {p:>7.3f} {stars}")

print(f"\n{'#'*55}")
print(f"  INFRASTRUCTURE WITH LAGS (t, t-1, t-2, t-3)")
print(f"{'#'*55}")
run_infra_lags(sub, 'primary_enroll_gross_pct',   'Primary Enrollment (%)')
run_infra_lags(sub, 'secondary_enroll_gross_pct', 'Secondary Enrollment (%)')
# ── Save infrastructure results to Excel ──
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = 'Infrastructure DiD'

BLUE  = PatternFill('solid', start_color='1F3D5C')
ALT   = PatternFill('solid', start_color='F5F9FC')
WHITE = PatternFill('solid', start_color='FFFFFF')
HFONT = Font(name='Arial', bold=True, color='FFFFFF', size=10)
BFONT = Font(name='Arial', size=10)
BDFONT= Font(name='Arial', bold=True, size=10)
TFONT = Font(name='Arial', bold=True, color='1F3D5C', size=12)
thin  = Side(style='thin', color='CCCCCC')
BDR   = Border(left=thin, right=thin, top=thin, bottom=thin)
CENTER= Alignment(horizontal='center')
LEFT  = Alignment(horizontal='left')
RIGHT = Alignment(horizontal='right')

def hdr(cell, text):
    cell.value, cell.font, cell.fill = text, HFONT, BLUE
    cell.alignment, cell.border = CENTER, BDR

def body(cell, val, fmt=None, bold=False, alt=False, align='right'):
    cell.value  = val
    cell.font   = BDFONT if bold else BFONT
    cell.fill   = ALT if alt else WHITE
    cell.border = BDR
    cell.alignment = RIGHT if align=='right' else LEFT if align=='left' else CENTER
    if fmt: cell.number_format = fmt

r = 1
ws.cell(r,1).value = 'Infrastructure Investment & Education — Dose-Response Results'
ws.cell(r,1).font  = TFONT
ws.merge_cells('A1:H1'); r+=1
ws.cell(r,1).value = '10 low-income countries | Energy + Transport investment from CGIT | Clustered SEs by country'
ws.cell(r,1).font  = Font(name='Arial', size=9, italic=True, color='666666')
ws.merge_cells('A2:H2'); r+=2

for j,h in enumerate(['Outcome','Specification','N Obs','Countries','Coefficient','Std Error','t-stat','p-value'],1):
    hdr(ws.cell(r,j), h)
r+=1

results = [
    ('Primary Enrollment (%)',   'Contemporaneous',         196, 10, -0.1730, 0.2340, -0.74, 0.460),
    ('Secondary Enrollment (%)', 'Contemporaneous',         155, 10, -0.0507, 0.3518, -0.14, 0.885),
    ('Primary Enrollment (%)',   'Log infra (t)',           170, 10, -0.0362, 0.2238, -0.16, 0.871),
    ('Primary Enrollment (%)',   'Log infra (t-1)',         170, 10, -0.0098, 0.2430, -0.04, 0.968),
    ('Primary Enrollment (%)',   'Log infra (t-2)',         170, 10, -0.0430, 0.1683, -0.26, 0.799),
    ('Primary Enrollment (%)',   'Log infra (t-3)',         170, 10, -0.2138, 0.1278, -1.67, 0.094),
    ('Secondary Enrollment (%)', 'Log infra (t)',           131, 10, +0.0189, 0.2580, +0.07, 0.942),
    ('Secondary Enrollment (%)', 'Log infra (t-1)',         131, 10, -0.0713, 0.2500, -0.29, 0.776),
    ('Secondary Enrollment (%)', 'Log infra (t-2)',         131, 10, -0.0856, 0.2333, -0.37, 0.714),
    ('Secondary Enrollment (%)', 'Log infra (t-3)',         131, 10, -0.2873, 0.2947, -0.97, 0.330),
]

for i,(olbl,spec,n,nc,coef,se,t,p) in enumerate(results):
    alt = i%2==0
    sig = '***' if p<0.01 else '**' if p<0.05 else '*' if p<0.1 else ''
    body(ws.cell(r,1), olbl, alt=alt, align='left')
    body(ws.cell(r,2), spec, alt=alt, align='left')
    body(ws.cell(r,3), n,    alt=alt, fmt='#,##0')
    body(ws.cell(r,4), nc,   alt=alt, fmt='#,##0')
    body(ws.cell(r,5), coef, alt=alt, fmt='+0.0000;-0.0000')
    body(ws.cell(r,6), se,   alt=alt, fmt='0.0000')
    body(ws.cell(r,7), t,    alt=alt, fmt='0.00')
    pval_str = f"{p:.3f} {sig}" if sig else f"{p:.3f}"
    body(ws.cell(r,8), pval_str, alt=alt, align='center', bold=(p<0.1))
    r+=1

r+=1
ws.cell(r,1).value = '*** p<0.01  ** p<0.05  * p<0.10  |  Countries: AFG, BEN, BGD, ETH, KHM, LAO, MOZ, SDN, TGO, TZA'
ws.cell(r,1).font  = Font(name='Arial', size=9, italic=True, color='888888')
ws.merge_cells(f'A{r}:H{r}')

for i,w in enumerate([25,22,10,10,13,13,10,12],1):
    ws.column_dimensions[get_column_letter(i)].width = w

wb.save('Data/Regressions/infrastructure_results.xlsx')
print("Saved: Data/Regressions/infrastructure_results.xlsx")